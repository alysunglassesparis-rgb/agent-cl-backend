from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.styles import PatternFill, Font
import io, zipfile, os, re, base64
from datetime import datetime, timedelta

app = Flask(__name__)
CORS(app, expose_headers=['X-Found-Optic','X-Found-Sun','X-Not-Found','X-Red-Refs'])

DATA_DIR = os.environ.get('RAILWAY_VOLUME_MOUNT_PATH', '/tmp/cl_data')
os.makedirs(DATA_DIR, exist_ok=True)
OPTIC_PATH = os.path.join(DATA_DIR, 'optic.xlsx')
SUN_PATH   = os.path.join(DATA_DIR, 'sun.xlsx')

CATALOGUE_OPTIC = None
CATALOGUE_SUN   = None
OPTIC_BYTES     = None
SUN_BYTES       = None
ADMIN_PASSWORD  = os.environ.get('ADMIN_PASSWORD', 'alysun2024')

NAME_KEYWORDS  = ['name', 'nom', 'reference', 'ref', 'style name']
QTY_KEYWORDS   = ['qty 1', 'qty1', 'quantity 1', 'qty', 'quantity']
PRICE_KEYWORDS = ['wholesale', 'prix', 'price', 'wholesale (eur)']

def normalize(s):
    return re.sub(r'\s+', ' ', str(s).strip()).lower()

def detect_columns(ws):
    cols = {'name': None, 'qty': None, 'wholesale': None}
    header_row = None
    for r in range(1, 5):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 30)]
        if any(v and any(k in normalize(str(v)) for k in ['name','qty','wholesale','season']) for v in row_vals):
            header_row = r
            break
    if not header_row:
        return cols, 2
    for c in range(1, 30):
        val = ws.cell(row=header_row, column=c).value
        if not val: continue
        v = normalize(str(val))
        if cols['name'] is None and v == 'name':
            cols['name'] = c
        if cols['qty'] is None:
            if v in ('qty 1', 'qty1', 'quantity 1'):
                cols['qty'] = c
            elif v in ('qty', 'quantity') and cols['qty'] is None:
                cols['qty'] = c
        if cols['wholesale'] is None:
            if any(k in v for k in PRICE_KEYWORDS):
                if 'total' not in v and 'msrp' not in v and 'size price' not in v:
                    cols['wholesale'] = c
    if cols['name'] is None:
        for c in range(1, 30):
            cell = ws.cell(row=header_row + 1, column=c)
            if cell.value and re.match(r'^CL\d+', str(cell.value)):
                cols['name'] = c
                break
    return cols, header_row + 1

def serial_to_year_month(serial):
    epoch = datetime(1899, 12, 30)
    d = epoch + timedelta(days=int(serial))
    return d.year, d.month

def extract_image_row_mapping(z):
    row_to_img = {}
    drawing_files = sorted([f for f in z.namelist() if re.match(r'xl/drawings/drawing\d+\.xml$', f)])
    for drawing_file in drawing_files:
        try:
            rels_path = drawing_file.replace('xl/drawings/', 'xl/drawings/_rels/') + '.rels'
            drawing_xml = z.read(drawing_file).decode('utf-8')
            rels_xml = z.read(rels_path).decode('utf-8')
            rel_map = {}
            for m in re.finditer(r'Id="([^"]+)"[^>]+Target="([^"]+)"', rels_xml):
                rel_map[m.group(1)] = m.group(2).replace('../', 'xl/')
            anchors = re.findall(
                r'<xdr:oneCellAnchor>.*?<xdr:row>(\d+)</xdr:row>.*?r:embed="([^"]+)".*?</xdr:oneCellAnchor>',
                drawing_xml, re.DOTALL)
            for row_str, rid in anchors:
                row = int(row_str) + 1
                if row not in row_to_img:
                    row_to_img[row] = rel_map.get(rid, '')
        except: pass
    return row_to_img

def build_catalogue(xlsx_bytes):
    items = []
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as z:
        row_to_img = extract_image_row_mapping(z)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb['NuORDER Order Data'] if 'NuORDER Order Data' in wb.sheetnames else wb.active
        cols, data_start = detect_columns(ws)
        name_col  = cols['name']
        qty_col   = cols['qty']
        price_col = cols['wholesale']
        print(f"Detected cols — name:{name_col} qty:{qty_col} price:{price_col} start:{data_start}")
        if not name_col:
            print("ERROR: could not detect name column!")
            return items
        for i, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=data_start):
            raw_name = row[name_col - 1] if name_col else None
            if not raw_name or str(raw_name).strip() in ('', 'Name'): continue
            name = re.sub(r'\s+', ' ', str(raw_name).strip())
            style_m = re.search(r'CL(\d+)', name)
            style = style_m.group(1) if style_m else ''
            wholesale = row[price_col - 1] if price_col else None
            img_b64 = ''
            img_file = row_to_img.get(i, '')
            if img_file:
                try:
                    img_b64 = base64.b64encode(z.read(img_file)).decode('utf-8')
                except: pass
            items.append({
                'name': name, 'style': style, 'row': i,
                'img': img_b64, 'wholesale': wholesale,
                'qty_col': qty_col, 'category': '',
            })
    return items

def get_qty_col_letter(catalogue):
    if catalogue:
        for item in catalogue:
            if item.get('qty_col'):
                col = item['qty_col']
                result = ''
                while col:
                    col, rem = divmod(col - 1, 26)
                    result = chr(65 + rem) + result
                return result
    return 'T'

def patch_xlsx_quantities(xlsx_bytes, row_updates, qty_col_letter='T'):
    updates = {f"{qty_col_letter}{row}": qty for row, qty in row_updates.items()}
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), 'r') as zin:
        sheet_files = sorted([f for f in zin.namelist() if re.match(r'xl/worksheets/sheet\d+\.xml$', f)])
        target = sheet_files[0] if sheet_files else None
        with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                data = zin.read(item)
                if item == target and updates:
                    xml = data.decode('utf-8')
                    for coord, qty in updates.items():
                        qs = str(int(qty) if isinstance(qty, float) and qty == int(qty) else qty)
                        pat = rf'<c r="{coord}"([^>]*)>(.*?)</c>'
                        def repl(m, q=qs, c=coord):
                            attrs = re.sub(r'\s*t=["\'][^"\']*["\']', '', m.group(1))
                            return f'<c r="{c}"{attrs}><v>{q}</v></c>'
                        if re.search(pat, xml, re.DOTALL):
                            xml = re.sub(pat, repl, xml, flags=re.DOTALL)
                        else:
                            rn = re.search(r'\d+', coord).group()
                            rp = rf'(<row r="{rn}"[^>]*>)(.*?)(</row>)'
                            def ins(m, c=coord, q=qs):
                                return f'{m.group(1)}{m.group(2)}<c r="{c}"><v>{q}</v></c>{m.group(3)}'
                            xml = re.sub(rp, ins, xml, flags=re.DOTALL)
                    data = xml.encode('utf-8')
                if item == 'xl/workbook.xml':
                    wbxml = data.decode('utf-8')
                    wbxml = re.sub(r'<calcPr[^/]*/>', '<calcPr fullCalcOnLoad="1"/>', wbxml)
                    if '<calcPr' not in wbxml:
                        wbxml = wbxml.replace('</workbook>', '<calcPr fullCalcOnLoad="1"/></workbook>')
                    data = wbxml.encode('utf-8')
                zout.writestr(item, data)
    return output.getvalue()

def load_from_disk():
    global CATALOGUE_OPTIC, CATALOGUE_SUN, OPTIC_BYTES, SUN_BYTES
    if os.path.exists(OPTIC_PATH):
        try:
            with open(OPTIC_PATH, 'rb') as f: OPTIC_BYTES = f.read()
            CATALOGUE_OPTIC = build_catalogue(OPTIC_BYTES)
            print(f"Loaded optic: {len(CATALOGUE_OPTIC)} refs")
        except Exception as e: print(f"Error loading optic: {e}")
    if os.path.exists(SUN_PATH):
        try:
            with open(SUN_PATH, 'rb') as f: SUN_BYTES = f.read()
            CATALOGUE_SUN = build_catalogue(SUN_BYTES)
            print(f"Loaded sun: {len(CATALOGUE_SUN)} refs")
        except Exception as e: print(f"Error loading sun: {e}")

load_from_disk()

@app.route('/health', methods=['GET'])
def health():
    return jsonify({
        "status": "ok",
        "optic_loaded": CATALOGUE_OPTIC is not None,
        "sun_loaded": CATALOGUE_SUN is not None,
        "optic_count": len(CATALOGUE_OPTIC) if CATALOGUE_OPTIC else 0,
        "sun_count": len(CATALOGUE_SUN) if CATALOGUE_SUN else 0,
        "optic_on_disk": os.path.exists(OPTIC_PATH),
        "sun_on_disk": os.path.exists(SUN_PATH),
    })

@app.route('/admin/upload', methods=['POST'])
def admin_upload():
    global CATALOGUE_OPTIC, CATALOGUE_SUN, OPTIC_BYTES, SUN_BYTES
    pwd = request.form.get('password', '')
    if pwd != ADMIN_PASSWORD:
        return jsonify({"error": "Mot de passe incorrect"}), 401
    updated = []
    if 'optique' in request.files:
        OPTIC_BYTES = request.files['optique'].read()
        with open(OPTIC_PATH, 'wb') as f: f.write(OPTIC_BYTES)
        CATALOGUE_OPTIC = build_catalogue(OPTIC_BYTES)
        updated.append(f"Optique: {len(CATALOGUE_OPTIC)} refs")
    if 'solaire' in request.files:
        SUN_BYTES = request.files['solaire'].read()
        with open(SUN_PATH, 'wb') as f: f.write(SUN_BYTES)
        CATALOGUE_SUN = build_catalogue(SUN_BYTES)
        updated.append(f"Solaire: {len(CATALOGUE_SUN)} refs")
    return jsonify({"success": True, "updated": updated})

@app.route('/catalogue', methods=['GET'])
def get_catalogue():
    optic = [{'name': i['name'], 'style': i['style'], 'row': i['row'], 'wholesale': i['wholesale'], 'category': i['category']} for i in (CATALOGUE_OPTIC or [])]
    sun   = [{'name': i['name'], 'style': i['style'], 'row': i['row'], 'wholesale': i['wholesale'], 'category': i['category']} for i in (CATALOGUE_SUN or [])]
    return jsonify({"optic": optic, "sun": sun})

@app.route('/image/<source>/<int:row>', methods=['GET'])
def get_image(source, row):
    cat = CATALOGUE_OPTIC if source == 'optic' else CATALOGUE_SUN
    if not cat: return jsonify({"error": "Catalogue non chargé"}), 404
    for item in cat:
        if item['row'] == row:
            return jsonify({"img": item['img'], "name": item['name']})
    return jsonify({"error": "Introuvable"}), 404

@app.route('/generate', methods=['POST'])
def generate():
    global OPTIC_BYTES, SUN_BYTES, CATALOGUE_OPTIC, CATALOGUE_SUN
    if not OPTIC_BYTES or not SUN_BYTES:
        return jsonify({"error": "Catalogues non chargés"}), 400
    data = request.json
    order = data.get('order', [])
    updates_optic = {}
    updates_sun = {}
    lookup_optic = {re.sub(r'\s+', ' ', i['name']).upper(): i['row'] for i in (CATALOGUE_OPTIC or [])}
    lookup_sun   = {re.sub(r'\s+', ' ', i['name']).upper(): i['row'] for i in (CATALOGUE_SUN or [])}
    for item in order:
        name = re.sub(r'\s+', ' ', item['name']).upper()
        qty = item['qty']
        if item['source'] == 'optic' and name in lookup_optic:
            updates_optic[lookup_optic[name]] = qty
        elif item['source'] == 'sun' and name in lookup_sun:
            updates_sun[lookup_sun[name]] = qty
    qty_col_optic = get_qty_col_letter(CATALOGUE_OPTIC)
    qty_col_sun   = get_qty_col_letter(CATALOGUE_SUN)
    patched_optic = patch_xlsx_quantities(OPTIC_BYTES, updates_optic, qty_col_optic)
    patched_sun   = patch_xlsx_quantities(SUN_BYTES, updates_sun, qty_col_sun)
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('commande_optique.xlsx', patched_optic)
        zf.writestr('commande_solaire.xlsx', patched_sun)
    zip_buf.seek(0)
    response = send_file(zip_buf, mimetype='application/zip',
                         as_attachment=True, download_name='commande_CL.zip')
    response.headers['X-Optic-Count'] = str(len(updates_optic))
    response.headers['X-Sun-Count'] = str(len(updates_sun))
    return response

@app.route('/process', methods=['POST'])
def process():
    try:
        f1 = request.files.get('commande')
        f2 = request.files.get('optique')
        f3 = request.files.get('solaire')
        if not f1 or not f2 or not f3:
            return jsonify({"error": "3 fichiers requis"}), 400
        bytes1 = f1.read(); bytes2 = f2.read(); bytes3 = f3.read()
        wb1 = openpyxl.load_workbook(io.BytesIO(bytes1), keep_vba=False)
        ws1 = wb1.active
        cat2 = build_catalogue(bytes2)
        cat3 = build_catalogue(bytes3)
        lookup2 = {re.sub(r'\s+', ' ', i['name']).upper(): i['row'] for i in cat2}
        lookup3 = {re.sub(r'\s+', ' ', i['name']).upper(): i['row'] for i in cat3}
        def optic_cands(s, m):
            c = str(m).zfill(2)
            return [f"CL{s} OPT {c}", f"CL{s} OPT  {c}"]
        def sun_cands(s, m):
            c = str(m).zfill(2)
            return [f"CL{s} SG OPT {c}", f"CL{s} SG {c}", f"CL{s} SG Z OPT {c}"]
        data_start = 11
        for i, row in enumerate(ws1.iter_rows(min_row=1, max_row=20), start=1):
            if row[0].value and 'REFERENCE' in str(row[0].value).upper():
                data_start = i + 1; break
        RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        RED_FONT = Font(color="CC0000", bold=True)
        found_optic = found_sun = not_found = 0
        red_refs = []
        updates2 = {}; updates3 = {}
        for row in ws1.iter_rows(min_row=data_start):
            ca, cb = row[0], row[1]
            if ca.value is None: continue
            style = month = None
            if hasattr(ca.value, 'year'): style = str(ca.value.year); month = ca.value.month
            elif isinstance(ca.value, (int, float)) and ca.value > 1000:
                y, m = serial_to_year_month(ca.value); style = str(y); month = m
            elif isinstance(ca.value, str):
                m = re.match(r'^(\d{3,5})[^0-9](\d{1,2})$', ca.value.strip())
                if m: style = m.group(1); month = int(m.group(2))
            if not style: continue
            qty = cb.value if cb.value not in (None, '', 0) else 1
            found = False
            for c in optic_cands(style, month):
                if c.upper() in lookup2: updates2[lookup2[c.upper()]] = qty; found_optic += 1; found = True; break
            if not found:
                for c in sun_cands(style, month):
                    if c.upper() in lookup3: updates3[lookup3[c.upper()]] = qty; found_sun += 1; found = True; break
            if not found:
                red_refs.append(f"{style}-{month}"); not_found += 1
                for cell in row:
                    cell.fill = RED_FILL
                    cell.font = Font(color="CC0000", bold=cell.font.bold if cell.font else False,
                                     size=cell.font.size if cell.font else None, name=cell.font.name if cell.font else None)
                row[2].value = '⚠ INTROUVABLE'; row[2].font = RED_FONT
        if not_found > 0:
            if '⚠ Refs introuvables' in wb1.sheetnames: del wb1['⚠ Refs introuvables']
            ws_err = wb1.create_sheet('⚠ Refs introuvables')
            ws_err['A1'] = 'Référence'; ws_err['B1'] = 'Statut'
            ws_err['A1'].font = Font(bold=True); ws_err['B1'].font = Font(bold=True)
            for i, ref in enumerate(red_refs, start=2):
                ws_err[f'A{i}'] = ref; ws_err[f'A{i}'].font = Font(color="CC0000", bold=True)
                ws_err[f'B{i}'] = '⚠ Introuvable'; ws_err[f'B{i}'].font = Font(color="CC0000")
        buf1 = io.BytesIO(); wb1.save(buf1); buf1.seek(0)
        qty2 = get_qty_col_letter(cat2)
        qty3 = get_qty_col_letter(cat3)
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f1.filename.replace('.xlsx','') + '_traite.xlsx', buf1.read())
            zf.writestr(f2.filename.replace('.xlsx','') + '_MAJ.xlsx', patch_xlsx_quantities(bytes2, updates2, qty2))
            zf.writestr(f3.filename.replace('.xlsx','') + '_MAJ.xlsx', patch_xlsx_quantities(bytes3, updates3, qty3))
        zip_buf.seek(0)
        response = send_file(zip_buf, mimetype='application/zip', as_attachment=True, download_name='commande_traitee.zip')
        response.headers['X-Found-Optic'] = str(found_optic)
        response.headers['X-Found-Sun'] = str(found_sun)
        response.headers['X-Not-Found'] = str(not_found)
        response.headers['X-Red-Refs'] = ','.join(red_refs)
        return response
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route('/sirene', methods=['GET'])
def search_sirene():
    q = request.args.get('q', '').strip()
    if not q: return jsonify({"error": "Query required"}), 400
    import urllib.request, urllib.parse, json as json_mod
    try:
        encoded = urllib.parse.quote(q)
        url = f"https://recherche-entreprises.api.gouv.fr/search?q={encoded}&limite=5"
        req = urllib.request.Request(url, headers={'User-Agent': 'AlysunGlasses/1.0'})
        with urllib.request.urlopen(req, timeout=5) as r:
            data = json_mod.loads(r.read().decode('utf-8'))
        results = []
        for r in data.get('results', []):
            siege = r.get('siege', {})
            siret = siege.get('siret', '')
            siren = r.get('siren', '')
            try:
                tva = f"FR{(12 + 3 * (int(siren) % 97)) % 97:02d}{siren}" if siren else ''
            except:
                tva = ''
            adresse_parts = [siege.get('numero_voie',''), siege.get('type_voie',''), siege.get('libelle_voie',''), siege.get('code_postal',''), siege.get('libelle_commune','')]
            adresse = ' '.join(p for p in adresse_parts if p).strip()
            results.append({'nom': r.get('nom_complet', ''), 'siret': siret, 'tva': tva, 'adresse': adresse})
        return jsonify({"results": results})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
